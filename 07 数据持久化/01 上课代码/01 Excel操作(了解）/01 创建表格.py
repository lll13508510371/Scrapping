# pip install openpyxl
import openpyxl  # 第三方模块

# 1.创建一个工作簿对象
work_book = openpyxl.Workbook()
print(work_book)

# 2.创建表格对象
sheet1 = work_book.create_sheet('表1')
print(sheet1)

# 操作表格...
# sheet1['C4'] = 1111
# sheet1['A2'] = 2222

# cell 单元格对象, row表示行, column表示列,
sheet1.cell(row=1, column=1).value = '666666'
sheet1.cell(row=2, column=2).value = '999999'

# 将列表或者元组中的数据整行写入到表格中
for_data = [1, 2, 3, 4, 5]
sheet1.append(for_data)
sheet1.append((6, 7, 8, 9, 10))

# 3.保存工作簿对象
work_book.save('实例.xlsx')

'''
如果是文件已经创建，往里面加数据的时候需要关闭文件，就像打开xlsx的时候改不了名一样会报错
'''
'''
工作簿对象没有保存之前是能够被覆盖的
'''