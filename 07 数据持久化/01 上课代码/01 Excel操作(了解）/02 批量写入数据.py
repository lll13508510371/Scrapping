import openpyxl

work = openpyxl.Workbook()
# sheet2 = work.create_sheet('表2')
sheet2 = work.active  # 使用默认表


for i in range(1, 10):
    for j in range(1, i + 1):
        print(f'{j} x {i} = {j * i}', end='\t')
        sheet2.cell(row=i, column=j).value = f'{j} x {i} = {j * i}'
    print()


work.save('九九乘法口诀表.xlsx')

# office  --> gbk
# wps  --> utf-8
