import openpyxl

workbook = openpyxl.load_workbook('九九乘法口诀表.xlsx')

# workbook.sheetnames  获取工作簿对象所有表的名字
print(workbook.sheetnames)

# 根据名字取出来表格
sheet = workbook['Sheet']

print(sheet.max_row)  # 最大行
print(sheet.max_column)  # 最大列

# # 获取第一行数据
# for i in range(1, sheet.max_column + 1):
#     print(sheet.cell(row=1, column=i).value)

# # 获取第一列数据
# for j in range(1, sheet.max_row + 1):
#     print(sheet.cell(row=j, column=1).value)

for i in range(1, sheet.max_column + 1):
    for j in range(1, i + 1):
        print(sheet.cell(row=i, column=j).value, end='\t')
    print()
